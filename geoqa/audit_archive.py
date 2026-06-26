from __future__ import annotations

import json
import re
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from geoqa.conversion import ArchiveSafetyLimits, _scan_zip_archive, list_vector_layers
from geoqa.execution import ValidationExecutionResult, validate_dataset_with_profile
from geoqa.problem_registry import VALIDATION_RULE_VERSION
from geoqa.reports.explanations import enrich_issue_row, explanation_for
from geoqa.reports.report_generator import generate_report, summarize_issues
from geoqa.validations.base import ValidationIssue


POINT_TERMS = ("meter", "valve", "hydrant", "chamber", "asset", "customer", "leak", "sensor", "flow_meter", "flow-meter")
LINE_TERMS = ("pipe", "network", "main", "service", "line", "route", "road")
POLYGON_TERMS = ("dma", "boundary", "zone", "parcel", "aoi", "district", "admin")


@dataclass(slots=True)
class LayerInventory:
    archive_name: str
    layer_id: str
    layer_path: str
    layer_name: str
    source_path: str
    ogr_layer: str | None = None
    geometry_type: str | None = None
    feature_count: int | None = None
    source_crs: str | None = None
    bbox: str | None = None
    attribute_fields: list[str] = field(default_factory=list)
    recommended_profile: str = "generic_quick"
    profile_reason: str = "Generic fallback profile."
    validated_yes_no: str = "no"
    validation_status: str = "not_run"
    issue_count: int = 0


@dataclass(slots=True)
class AuditArchiveResult:
    input_path: str
    output_dir: str
    archive_name: str
    layers: list[LayerInventory]
    validated_results: list[ValidationExecutionResult]
    json_reports: list[str]
    excel_report: str | None
    summary_report: str
    pdf_report: str | None
    command: str


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    cleaned = cleaned.strip("._")
    return cleaned or "layer"


def _safe_crs(layer: Any) -> str | None:
    crs = getattr(layer, "crs", None)
    if crs is None:
        return None
    try:
        return crs.to_string()
    except Exception:
        return str(crs)


def _safe_bbox(layer: Any) -> str | None:
    try:
        min_x, min_y, max_x, max_y = layer.total_bounds
        return f"{min_x:.6f}, {min_y:.6f}, {max_x:.6f}, {max_y:.6f}"
    except Exception:
        return None


def _geometry_type(layer: Any) -> str | None:
    try:
        counts = layer.geometry.geom_type.value_counts().to_dict()
        if not counts:
            return None
        return str(sorted(counts.items(), key=lambda item: (-int(item[1]), str(item[0])))[0][0])
    except Exception:
        return None


def _load_layer(path: Path, *, ogr_layer: str | None = None) -> Any:
    try:
        import geopandas as gpd
    except ImportError as exc:
        raise RuntimeError("GeoQA archive audit requires GeoPandas.") from exc
    if ogr_layer:
        return gpd.read_file(path, layer=ogr_layer)
    return gpd.read_file(path)


def _recommend_profile(layer_name: str, geometry_type: str | None) -> tuple[str, str]:
    normalized_name = layer_name.lower()
    normalized_geometry = (geometry_type or "").lower()
    if normalized_geometry in {"point", "multipoint"}:
        if any(term in normalized_name for term in POINT_TERMS):
            return "point_asset_quick", "Point asset layer name matched asset delivery terms."
        return "generic_quick", "Point layer without a specialized asset name uses generic quick QA."
    if normalized_geometry in {"linestring", "multilinestring"}:
        if any(term in normalized_name for term in LINE_TERMS):
            return "water_network_quick", "Line layer name matched network or pipe delivery terms."
        return "line_network_quick", "Line layer uses line network quick QA."
    if normalized_geometry in {"polygon", "multipolygon"}:
        if any(term in normalized_name for term in POLYGON_TERMS):
            return "boundaries_quick", "Polygon layer name matched boundary, parcel, zone, or admin terms."
        return "generic_quick", "Polygon layer without a specialized name uses generic quick QA."
    return "generic_quick", "Unknown geometry type uses generic quick QA."


def _existing_profile_name(profile_name: str) -> str:
    from geoqa.profile_registry import get_geoqa_profile

    if get_geoqa_profile(profile_name) is not None:
        return profile_name
    fallbacks = {
        "point_asset_quick": "generic_quick",
        "line_network_quick": "generic_quick",
    }
    return fallbacks.get(profile_name, "generic_quick")


def _inventory_layer(
    *,
    archive_name: str,
    layer_path: str,
    layer_name: str,
    source_path: Path,
    ogr_layer: str | None = None,
) -> LayerInventory:
    layer = _load_layer(source_path, ogr_layer=ogr_layer)
    geometry_type = _geometry_type(layer)
    recommended_profile, profile_reason = _recommend_profile(layer_name, geometry_type)
    return LayerInventory(
        archive_name=archive_name,
        layer_id=_slug(layer_path),
        layer_path=layer_path,
        layer_name=layer_name,
        source_path=str(source_path),
        ogr_layer=ogr_layer,
        geometry_type=geometry_type,
        feature_count=len(layer) if hasattr(layer, "__len__") else None,
        source_crs=_safe_crs(layer),
        bbox=_safe_bbox(layer),
        attribute_fields=[str(column) for column in getattr(layer, "columns", []) if str(column) != "geometry"],
        recommended_profile=recommended_profile,
        profile_reason=profile_reason,
    )


def discover_layers(path: str | Path, *, temp_root: Path | None = None) -> tuple[list[LayerInventory], Path | None]:
    input_path = Path(path).resolve()
    archive_name = input_path.name
    extracted_root: Path | None = None

    if input_path.suffix.lower() == ".zip":
        _scan_zip_archive(input_path, ArchiveSafetyLimits(max_member_count=1000, max_uncompressed_size_mb=2048.0))
        extracted_root = Path(tempfile.mkdtemp(prefix="geoqa_audit_archive_", dir=temp_root))
        with ZipFile(input_path) as archive:
            archive.extractall(extracted_root)
        layers = []
        for shp_path in sorted(extracted_root.rglob("*.shp")):
            relative = shp_path.relative_to(extracted_root).as_posix()
            layers.append(
                _inventory_layer(
                    archive_name=archive_name,
                    layer_path=Path(relative).with_suffix("").as_posix(),
                    layer_name=shp_path.stem,
                    source_path=shp_path,
                )
            )
        return layers, extracted_root

    if input_path.is_dir():
        layers = []
        for shp_path in sorted(input_path.rglob("*.shp")):
            relative = shp_path.relative_to(input_path).as_posix()
            layers.append(
                _inventory_layer(
                    archive_name=archive_name,
                    layer_path=Path(relative).with_suffix("").as_posix(),
                    layer_name=shp_path.stem,
                    source_path=shp_path,
                )
            )
        for file_path in sorted(input_path.rglob("*")):
            if file_path.suffix.lower() in {".geojson", ".json"}:
                relative = file_path.relative_to(input_path).as_posix()
                layers.append(
                    _inventory_layer(
                        archive_name=archive_name,
                        layer_path=relative,
                        layer_name=file_path.stem,
                        source_path=file_path,
                    )
                )
        return layers, None

    if input_path.suffix.lower() == ".gpkg":
        layers = []
        names = list_vector_layers(input_path)
        for name in names:
            layers.append(
                _inventory_layer(
                    archive_name=archive_name,
                    layer_path=name,
                    layer_name=name,
                    source_path=input_path,
                    ogr_layer=name,
                )
            )
        return layers, None

    return [
        _inventory_layer(
            archive_name=archive_name,
            layer_path=input_path.name,
            layer_name=input_path.stem,
            source_path=input_path,
        )
    ], None


def _selected_layers(layers: list[LayerInventory], *, selected_layer: str | None, all_layers: bool) -> list[LayerInventory]:
    if selected_layer:
        selected = [layer for layer in layers if selected_layer in {layer.layer_path, layer.layer_name, layer.layer_id}]
        if not selected:
            raise ValueError(f"Selected layer was not found: {selected_layer}")
        return selected
    if all_layers:
        return layers
    return layers


def _issue_rows(result: ValidationExecutionResult, layer: LayerInventory, *, no_coordinates: bool) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for issue in result.issues:
        row = enrich_issue_row(issue.to_dict())
        row["layer_name"] = layer.layer_name
        row["layer_path"] = layer.layer_path
        row["geometry_type"] = layer.geometry_type
        geometry = row.get("geometry")
        if no_coordinates:
            row["geometry"] = None
            row["issue_location_wkt"] = None
            row["x_coordinate"] = None
            row["y_coordinate"] = None
        else:
            row["issue_location_wkt"] = geometry if isinstance(geometry, str) else None
            row["x_coordinate"] = None
            row["y_coordinate"] = None
        rows.append(row)
    return rows


def _plain_summary(
    *,
    archive_name: str,
    layers: list[LayerInventory],
    validated: list[LayerInventory],
    issue_count: int,
) -> str:
    if validated:
        validated_text = ", ".join(layer.layer_name for layer in validated[:3])
        if len(validated) > 3:
            validated_text += f" and {len(validated) - 3} more"
    else:
        validated_text = "no layers"
    return (
        f"GeoQA reviewed {len(layers)} detected layer(s) in {archive_name} and validated {len(validated)} layer(s): "
        f"{validated_text}. GeoQA found {issue_count} issue(s). Source layers were not modified. "
        "Validator coverage records show which checks completed or were skipped because the geometry type was not applicable."
    )


def _write_summary_text(path: Path, text: str) -> Path:
    path.write_text(text + "\n", encoding="utf-8")
    return path


def _write_summary_pdf(path: Path, *, title: str, summary: str, layers: list[LayerInventory]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [title, "", summary, "", "Layer inventory"]
    for layer in layers[:25]:
        lines.append(f"{layer.layer_name}  {layer.geometry_type or 'unknown'}  {layer.feature_count or 0} features  {layer.validation_status}")
    lines.append("")
    lines.append("Generated by GeoQA")

    def escape_pdf_text(value: str) -> str:
        return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    content_lines = ["BT", "/F1 11 Tf", "72 740 Td", "14 TL"]
    for raw_line in lines:
        for start in range(0, max(len(raw_line), 1), 95):
            line = raw_line[start : start + 95]
            if not line and start > 0:
                continue
            content_lines.append(f"({escape_pdf_text(line)}) Tj")
            content_lines.append("T*")
            if start + 95 >= len(raw_line):
                break
    content_lines.append("ET")
    content = "\n".join(content_lines).encode("latin-1", errors="replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(bytes(pdf))
    return path


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _append_dict_rows(sheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_cell_value(row.get(header)) for header in headers])


def _style_workbook(workbook: Any) -> None:
    from openpyxl.styles import Font, PatternFill

    severity_fills = {
        "critical": PatternFill("solid", fgColor="7F1D1D"),
        "high": PatternFill("solid", fgColor="C2410C"),
        "medium": PatternFill("solid", fgColor="FACC15"),
        "low": PatternFill("solid", fgColor="BFDBFE"),
        "informational": PatternFill("solid", fgColor="D1D5DB"),
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        header_map = {str(cell.value): cell.column for cell in sheet[1]}
        severity_column = header_map.get("severity")
        if severity_column:
            for row in sheet.iter_rows(min_row=2):
                value = str(row[severity_column - 1].value or "").lower()
                fill = severity_fills.get(value)
                if fill:
                    row[severity_column - 1].fill = fill


def write_excel_report(
    *,
    output_path: str | Path,
    archive_name: str,
    input_path: str,
    output_dir: str,
    command: str,
    layers: list[LayerInventory],
    validated_layers: list[LayerInventory],
    results: list[ValidationExecutionResult],
    issue_rows: list[dict[str, Any]],
    plain_summary: str,
    expected_crs: str | None,
    sanitize: bool,
) -> Path:
    from openpyxl import Workbook

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"

    total_features = sum(int(layer.feature_count or 0) for layer in validated_layers)
    severities: dict[str, int] = {}
    for row in issue_rows:
        severity = str(row.get("severity") or "informational").lower()
        severities[severity] = severities.get(severity, 0) + 1
    summary_rows = [
        ("GeoQA", "Generated by GeoQA"),
        ("Report title", "GeoQA Audit Report"),
        ("Dataset/archive name", archive_name),
        ("Generated date/time", datetime.now().isoformat(timespec="seconds")),
        ("Validation rule version", VALIDATION_RULE_VERSION),
        ("Total layers detected", len(layers)),
        ("Total layers validated", len(validated_layers)),
        ("Total features checked", total_features),
        ("Total issues", len(issue_rows)),
        ("Critical issues", severities.get("critical", 0)),
        ("High issues", severities.get("high", 0)),
        ("Medium issues", severities.get("medium", 0)),
        ("Low/informational issues", severities.get("low", 0) + severities.get("informational", 0)),
        ("Execution status", "complete" if all(result.completed for result in results) else "partial"),
        ("Plain-English summary", plain_summary),
    ]
    for row in summary_rows:
        summary_sheet.append(row)

    inventory_sheet = workbook.create_sheet("Layer Inventory")
    inventory_headers = [
        "archive_name",
        "layer_id",
        "layer_path",
        "layer_name",
        "geometry_type",
        "feature_count",
        "source_crs",
        "bbox",
        "recommended_profile",
        "profile_reason",
        "validated_yes_no",
        "validation_status",
        "issue_count",
    ]
    _append_dict_rows(inventory_sheet, inventory_headers, [asdict(layer) for layer in layers])

    issue_summary_sheet = workbook.create_sheet("Issue Summary")
    summary_by_problem: dict[str, dict[str, Any]] = {}
    for row in issue_rows:
        problem = str(row.get("problem_name"))
        item = summary_by_problem.setdefault(
            problem,
            {
                "issue_type": problem,
                "severity": row.get("severity"),
                "confidence": row.get("confidence"),
                "priority_score": row.get("priority_score"),
                "affected_features_count": 0,
                "issue_class": row.get("issue_class"),
                "iso_category": row.get("iso_category"),
                "root_cause": (row.get("provenance") or {}).get("catalog_category") if isinstance(row.get("provenance"), dict) else None,
                "plain_english_explanation": row.get("what_geoqa_found"),
                "recommended_action": row.get("recommended_action"),
            },
        )
        item["affected_features_count"] += 1
    _append_dict_rows(
        issue_summary_sheet,
        [
            "issue_type",
            "severity",
            "confidence",
            "priority_score",
            "affected_features_count",
            "issue_class",
            "iso_category",
            "root_cause",
            "plain_english_explanation",
            "recommended_action",
        ],
        list(summary_by_problem.values()),
    )

    row_sheet = workbook.create_sheet("Row-Level Issues")
    _append_dict_rows(
        row_sheet,
        [
            "layer_name",
            "layer_path",
            "feature_id",
            "source_row_index",
            "asset_id",
            "problem_name",
            "severity",
            "confidence",
            "priority_score",
            "geometry_type",
            "x_coordinate",
            "y_coordinate",
            "issue_location_wkt",
            "related_feature_ids",
            "description",
            "why_it_matters",
            "recommended_action",
            "review_status",
            "reviewer_notes",
        ],
        issue_rows,
    )

    coverage_sheet = workbook.create_sheet("Validator Coverage")
    coverage_rows = []
    for result in results:
        coverage_rows.extend(result.validator_coverage or result.summary.get("validator_coverage", []))
    _append_dict_rows(
        coverage_sheet,
        ["layer_name", "validator_name", "status", "reason", "geometry_type", "expected_geometry_types", "profile", "notes"],
        coverage_rows,
    )

    crs_sheet = workbook.create_sheet("CRS Review")
    crs_rows = []
    for layer in layers:
        crs_status = "missing" if not layer.source_crs else "valid"
        action_required = "Confirm CRS before analysis." if not layer.source_crs else "No CRS action required unless an expected CRS is configured."
        if expected_crs and layer.source_crs and layer.source_crs != expected_crs:
            crs_status = "expected_crs_mismatch"
            action_required = "Confirm the authoritative CRS before redefining or reprojecting."
        crs_rows.append(
            {
                "layer_name": layer.layer_name,
                "source_crs": layer.source_crs,
                "expected_crs": expected_crs,
                "display_crs": "EPSG:4326",
                "crs_status": crs_status,
                "transformation_available": bool(layer.source_crs),
                "action_required": action_required,
                "notes": "Display reprojection may be required for web maps. Source CRS is not modified.",
            }
        )
    _append_dict_rows(
        crs_sheet,
        ["layer_name", "source_crs", "expected_crs", "display_crs", "crs_status", "transformation_available", "action_required", "notes"],
        crs_rows,
    )

    fix_sheet = workbook.create_sheet("Fix Plan")
    _append_dict_rows(
        fix_sheet,
        [
            "issue_id",
            "layer_name",
            "feature_id",
            "problem_name",
            "suggested_fix",
            "safe_to_auto_fix",
            "requires_client_review",
            "approved_by",
            "approval_date",
            "applied",
            "applied_output_file",
        ],
        [
            {
                "issue_id": row.get("issue_id"),
                "layer_name": row.get("layer_name"),
                "feature_id": row.get("feature_id"),
                "problem_name": row.get("problem_name"),
                "suggested_fix": row.get("recommended_action"),
                "safe_to_auto_fix": "no",
                "requires_client_review": "yes",
                "approved_by": "",
                "approval_date": "",
                "applied": "no",
                "applied_output_file": "",
            }
            for row in issue_rows
        ],
    )

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_rows = [
        ("GeoQA version", "local"),
        ("validation_rule_version", VALIDATION_RULE_VERSION),
        ("run command", command),
        ("input path", input_path if not sanitize else Path(input_path).name),
        ("output path", output_dir if not sanitize else Path(output_dir).name),
        ("timestamp", datetime.now().isoformat(timespec="seconds")),
        ("profiles used", ", ".join(sorted({layer.recommended_profile for layer in validated_layers}))),
        ("expected CRS configuration", expected_crs or ""),
        ("sanitize mode", str(sanitize)),
        ("limitations", "Audit only. Source data is not modified by default."),
    ]
    for row in metadata_rows:
        metadata_sheet.append(row)

    _style_workbook(workbook)
    workbook.save(output)
    return output


def run_audit_archive(
    path: str | Path,
    *,
    output_dir: str | Path,
    profiles: str = "auto",
    excel: bool = True,
    json_reports: bool = True,
    selected_layer: str | None = None,
    all_layers: bool = False,
    expected_crs: str | None = None,
    sanitize: bool = False,
    no_coordinates: bool = False,
    public_demo_mode: bool = False,
    write_fix_plan: bool = False,
    apply_safe_fixes: bool = False,
    command: str = "geoqa audit-archive",
) -> AuditArchiveResult:
    if apply_safe_fixes:
        raise ValueError("--apply-safe-fixes is not implemented for archive audit. Default mode is audit-only.")
    output = Path(output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    input_path = Path(path).resolve()
    temp_root = Path(tempfile.mkdtemp(prefix="geoqa_audit_work_"))
    layers, extracted_root = discover_layers(input_path, temp_root=temp_root)
    selected = _selected_layers(layers, selected_layer=selected_layer, all_layers=all_layers)

    results: list[ValidationExecutionResult] = []
    report_paths: list[str] = []
    all_issue_rows: list[dict[str, Any]] = []
    for layer in selected:
        profile_name = _existing_profile_name(layer.recommended_profile if profiles == "auto" else profiles)
        result = validate_dataset_with_profile(
            layer.source_path,
            profile=profile_name,
            expected_crs=expected_crs,
            ogr_layer=layer.ogr_layer,
            output_format=None,
            report_path=None,
            thermal_profile="cool",
            prefer_high_priority=True,
        )
        layer.validated_yes_no = "yes"
        layer.validation_status = result.execution_status
        layer.issue_count = len(result.issues)
        result.summary["layer_inventory"] = asdict(layer)
        result.summary["plain_english_summary"] = _plain_summary(
            archive_name=input_path.name,
            layers=layers,
            validated=[layer],
            issue_count=len(result.issues),
        )
        results.append(result)
        all_issue_rows.extend(_issue_rows(result, layer, no_coordinates=no_coordinates))
        if json_reports:
            report_base = output / f"{layer.layer_id}_geoqa_report"
            report_path = generate_report(result.issues, output_format="json", file_path=str(report_base), summary=result.summary)
            report_paths.append(str(report_path))

    plain_summary = _plain_summary(
        archive_name=input_path.name,
        layers=layers,
        validated=selected,
        issue_count=sum(len(result.issues) for result in results),
    )
    summary_path = _write_summary_text(output / "GeoQA_Audit_Summary.txt", plain_summary)
    pdf_path = _write_summary_pdf(output / "GeoQA_Audit_Summary.pdf", title="GeoQA Audit Summary", summary=plain_summary, layers=layers)
    excel_path = None
    if excel:
        date_token = datetime.now().strftime("%Y%m%d")
        workbook_name = f"GeoQA_Audit_Report_{_slug(input_path.stem)}_{date_token}.xlsx"
        excel_path = write_excel_report(
            output_path=output / workbook_name,
            archive_name=input_path.name,
            input_path=str(input_path),
            output_dir=str(output),
            command=command,
            layers=layers,
            validated_layers=selected,
            results=results,
            issue_rows=all_issue_rows,
            plain_summary=plain_summary,
            expected_crs=expected_crs,
            sanitize=sanitize or public_demo_mode,
        )
    if write_fix_plan:
        (output / "GeoQA_Fix_Plan_README.txt").write_text(
            "GeoQA wrote a human review fix plan in the Excel workbook. No fixes were applied.\n",
            encoding="utf-8",
        )
    if extracted_root is not None:
        pass
    return AuditArchiveResult(
        input_path=str(input_path),
        output_dir=str(output),
        archive_name=input_path.name,
        layers=layers,
        validated_results=results,
        json_reports=report_paths,
        excel_report=str(excel_path) if excel_path else None,
        summary_report=str(summary_path),
        pdf_report=str(pdf_path),
        command=command,
    )


__all__ = [
    "AuditArchiveResult",
    "LayerInventory",
    "discover_layers",
    "run_audit_archive",
    "write_excel_report",
]
